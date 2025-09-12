# -*- coding: utf-8 -*-
"""
Excel file management web interface
"""
import os
import time
from datetime import datetime
import openpyxl
from flask import request, jsonify
import globals as g

def register_excel_routes(app):
    """Регистрирует маршруты для управления Excel файлами"""
    
    @app.route("/manage-excel", methods=["GET", "POST"])
    def manage_excel():
        """Управление Excel файлами через веб-интерфейс"""
        if request.method == "GET":
            # Возвращаем HTML интерфейс
            return """
            <!DOCTYPE html>
            <html>
            <head>
                <title>ElevenLabs Excel Manager</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 20px; }
                    .container { max-width: 800px; margin: 0 auto; }
                    .btn { padding: 10px 20px; margin: 5px; background: #007bff; color: white; border: none; cursor: pointer; }
                    .btn:hover { background: #0056b3; }
                    .result { margin: 20px 0; padding: 15px; background: #f8f9fa; border: 1px solid #dee2e6; }
                    .error { background: #f8d7da; border-color: #f5c6cb; }
                    .success { background: #d4edda; border-color: #c3e6cb; }
                    .stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; }
                    .stat-card { background: white; padding: 15px; border-radius: 5px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>🎤 ElevenLabs Excel Manager</h1>
                    
                    <div class="stats" id="stats">
                        <div class="stat-card">
                            <h3>📊 Загрузка статистики...</h3>
                        </div>
                    </div>
                    
                    <h2>🛠️ Управление файлами</h2>
                    <button class="btn" onclick="createApiExcel()">📝 Создать api_elevenlabs.xlsx</button>
                    <button class="btn" onclick="checkQuotas()">🔍 Проверить квоты</button>
                    <button class="btn" onclick="loadStats()">📊 Обновить статистику</button>
                    
                    <div id="result" class="result" style="display: none;"></div>
                    
                    <h2>📋 Активные API ключи</h2>
                    <div id="apiKeys"></div>
                </div>
                
                <script>
                    function showResult(message, isError = false) {
                        const result = document.getElementById('result');
                        result.innerHTML = message;
                        result.className = 'result ' + (isError ? 'error' : 'success');
                        result.style.display = 'block';
                    }
                    
                    function createApiExcel() {
                        fetch('/manage-excel', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ action: 'create_api_excel' })
                        })
                        .then(r => r.json())
                        .then(data => {
                            showResult(data.message, !data.success);
                            if (data.success) loadStats();
                        });
                    }
                    
                    function checkQuotas() {
                        showResult('🔍 Проверка квот... Это может занять несколько минут.');
                        fetch('/manage-excel', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ action: 'check_quotas' })
                        })
                        .then(r => r.json())
                        .then(data => {
                            showResult(data.message, !data.success);
                            if (data.success) loadStats();
                        });
                    }
                    
                    function loadStats() {
                        fetch('/elevenlabs-stats')
                        .then(r => r.json())
                        .then(data => {
                            document.getElementById('stats').innerHTML = `
                                <div class="stat-card">
                                    <h3>🔑 API Ключи</h3>
                                    <p>Всего: ${data.total_api_keys || 0}</p>
                                    <p>Активных: ${data.active_keys || 0}</p>
                                </div>
                                <div class="stat-card">
                                    <h3>📊 Квота</h3>
                                    <p>Всего: ${(data.total_quota || 0).toLocaleString()}</p>
                                    <p>Доступно: ${(data.remaining_quota || 0).toLocaleString()}</p>
                                </div>
                                `;

                            loadApiKeys();
                        });
                    }
                    
                    function loadApiKeys() {
                        fetch('/api-keys-list')
                        .then(r => r.json())
                        .then(data => {
                            const html = data.map(key => `
                                <div class="stat-card">
                                    <strong>${key.email}</strong><br>
                                    <small>Квота: ${key.quota_remaining.toLocaleString()} / ${key.quota_total.toLocaleString()}</small><br>
                                    <small>Статус: ${key.status}</small><br>
                                    <small>Последняя проверка: ${key.last_checked}</small>
                                </div>
                            `).join('');
                            document.getElementById('apiKeys').innerHTML = html;
                        });
                    }
                    
                    // Загружаем статистику при открытии страницы
                    loadStats();
                    
                    // Автообновление каждые 30 секунд
                    setInterval(loadStats, 30000);
                </script>
            </body>
            </html>
            """
        
        # POST запрос для выполнения действий
        try:
            data = request.get_json()
            action = data.get('action')
            
            if action == 'create_api_excel':
                success, message = create_api_excel_file()
                return jsonify({"success": success, "message": message})
                
            elif action == 'check_quotas':
                success, message = check_all_quotas()
                return jsonify({"success": success, "message": message})
                
            else:
                return jsonify({"success": False, "message": "Unknown action"})
                
        except Exception as e:
            return jsonify({"success": False, "message": f"Error: {str(e)}"})


    @app.route("/api-keys-list")
    def api_keys_list():
        """Возвращает список API ключей с информацией"""
        try:
            filename = "api_elevenlabs.xlsx"
            
            if not os.path.exists(filename):
                return jsonify([])
            
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            
            keys = []
            for row in range(2, ws.max_row + 1):
                api_key = ws.cell(row=row, column=1).value
                if not api_key:
                    continue
                    
                email = ws.cell(row=row, column=2).value or "Unknown"
                quota_remaining = ws.cell(row=row, column=4).value or 0
                last_checked = ws.cell(row=row, column=5).value or "Never"
                status = ws.cell(row=row, column=6).value or "Unknown"
                usage_count = ws.cell(row=row, column=7).value or 0
                total_used = ws.cell(row=row, column=8).value or 0
                
                keys.append({
                    "email": email,
                    "quota_remaining": quota_remaining,
                    "quota_total": quota_remaining + total_used,
                    "last_checked": last_checked,
                    "status": status,
                    "usage_count": usage_count,
                    "total_used": total_used
                })
            
            return jsonify(keys)
            
        except Exception as e:
            return jsonify({"error": str(e)})


def create_api_excel_file():
    """Создает Excel файл для ElevenLabs API ключей"""
    try:
        filename = "api_elevenlabs.xlsx"
        
        if os.path.exists(filename):
            return False, f"Файл {filename} уже существует"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ElevenLabs APIs"
        
        # Заголовки
        headers = ["API Key", "Email", "Password", "Quota Remaining", "Last Checked", "Status", "Usage Count", "Total Used This Month", "Notes"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Пример данных
        example_data = [
            ["sk_example_key_1", "user1@example.com", "password123", 10000, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "active", 0, 0, "Замените на реальные данные"],
            ["sk_example_key_2", "user2@example.com", "password456", 8500, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "active", 150, 1500, "Замените на реальные данные"]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(filename)
        return True, f"✅ Создан файл {filename}. Не забудьте заменить примеры на реальные API ключи!"
        
    except Exception as e:
        return False, f"❌ Ошибка создания файла: {str(e)}"



def check_all_quotas():
    """Проверяет квоту для всех ключей ElevenLabs"""
    try:
        filename = "api_elevenlabs.xlsx"
        
        if not os.path.exists(filename):
            return False, f"Файл {filename} не найден! Создайте его сначала."
        
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        updated_count = 0
        errors = []
        
        for row in range(2, ws.max_row + 1):
            api_key = ws.cell(row=row, column=1).value
            email = ws.cell(row=row, column=2).value
            status = ws.cell(row=row, column=6).value
            
            if not api_key or status == "disabled":
                continue
            
            try:
                # Получаем прокси для проверки
                proxy_obj = g.proxy_manager.get_available_proxy(for_openai_fm=False)
                proxy_dict = g.elevenlabs_manager._get_proxy_dict(proxy_obj) if proxy_obj else {}
                quota = g.elevenlabs_manager.check_quota(api_key, proxy_dict)
                # Обновляем данные в Excel
                ws.cell(row=row, column=4, value=quota)  # Quota Remaining
                ws.cell(row=row, column=5, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))  # Last Checked
                
                if quota <= 0:
                    ws.cell(row=row, column=6, value="exhausted")
                else:
                    ws.cell(row=row, column=6, value="active")
                
                updated_count += 1
                
                # Задержка между запросами
                time.sleep(1)
                
            except Exception as e:
                errors.append(f"{email}: {str(e)}")
                ws.cell(row=row, column=6, value="error")
        
        wb.save(filename)
        
        message = f"✅ Проверено {updated_count} API ключей"
        if errors:
            message += f"\n❌ Ошибки: {'; '.join(errors[:3])}"
            if len(errors) > 3:
                message += f" и еще {len(errors) - 3}"
        
        return True, message
        
    except Exception as e:
        return False, f"❌ Ошибка проверки квот: {str(e)}"
