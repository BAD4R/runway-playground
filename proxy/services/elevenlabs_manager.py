# -*- coding: utf-8 -*-
"""
ElevenLabs API management and queue processing
"""
import os
import time
import uuid
import threading
import json
from datetime import datetime
from pathlib import Path
import openpyxl
from zipfile import BadZipFile
import requests
import queue as thread_queue
import math
from typing import Dict
from utils.logger import log, FULL_LOGS, maybe_truncate
import globals as g
import random
import socket
import re
from requests.exceptions import ReadTimeout, ConnectionError
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from config.global_params import load_elevenlabs_limits

from proxy.mobile_proxy import MobileProxyManager

VOICE_DEFAULTS = {
    'stability': 0.5,
    'similarity_boost': 0.5,
    'style': 0.5,
    'speed': 1.0,
    'use_speaker_boost': True,
}

MODEL_VOICE_PARAMS = {
    'eleven_monolingual_v1': {'stability', 'similarity_boost', 'style', 'use_speaker_boost'},
    'eleven_monolingual_v2': {'stability', 'similarity_boost', 'style', 'speed', 'use_speaker_boost'},
    'eleven_multilingual_v1': {'stability', 'similarity_boost', 'style', 'use_speaker_boost'},
    'eleven_multilingual_v2': {'stability', 'similarity_boost', 'style', 'speed', 'use_speaker_boost'},
    'eleven_turbo_v2': {'stability', 'similarity_boost', 'style', 'speed', 'use_speaker_boost'},
    'eleven_turbo_v2_5': {'stability', 'similarity_boost', 'style', 'speed', 'use_speaker_boost'},
    'eleven_flash_v2': {'stability', 'similarity_boost', 'speed'},
    'eleven_flash_v2_5': {'stability', 'similarity_boost', 'speed'},
}

# Models that cost 50% quota (rounded up)
DISCOUNTED_MODELS = {
    'eleven_flash_v2_5',
    'eleven_turbo_v2_5',
    'eleven_turbo_v2',
    'eleven_flash_v2',
}

# Spisok bazovykh golosov ElevenLabs, kotorye NE nuzhno udalyat
BASIC_VOICE_IDS = {
    '9BWtsMINqrJLrRacOk9x',  # Aria
    'EXAVITQu4vr4xnSDxMaL',  # Sarah
    'FGY2WhTYpPnrIDTdsKH5',  # Laura
    'IKne3meq5aSn9XLyUdCD',  # Charlie
    'JBFqnCBsd6RMkjVDRZzb',  # George
    'N2lVS1w4EtoT3dr4eOWO',  # Callum
    'SAz9YHcvj6GT2YYXdXww',  # River
    'TX3LPaxmHKxFdv7VOQHJ',  # Liam
    'XB0fDUnXU5powFXDhCwa',  # Charlotte
    'Xb7hH8MSUJpSbSDYk0k2',  # Alice
    'XrExE9yKIg1WjnnlVkGX',  # Matilda
    'bIHbv24MWmeRgasZH58o',  # Will
    'cgSgspJ2msm6clMCkdW9',  # Jessica
    'cjVigY5qzO86Huf0OWal',  # Eric
    'iP95p4xoKVk53GoZ742B',  # Chris
    'nPczCjzI2devNBz1zQrb',  # Brian
    'onwK4e9ZLuTAKqWW03F9',  # Daniel
    'pFZP5JQG7iQjIQuC4Bku',  # Lily
    'pqHfZKP75CvOlQylNhV4',  # Bill
    '21m00Tcm4TlvDq8ikWAM',  # Rachel
    '2EiwWnXFnvU5JabPnv8n',  # Clyde
    'CwhRBWXzGAHq8TQ4Fs17',  # Roger
    'GBv7mTt0atIp3Br8iCZE',  # Thomas
    'SOYHLrjzK2X1ezoPC6cr',  # Harry
}

class ElevenLabsManager:
    def __init__(self, excel_path: str = "api_elevenlabs.xlsx"):
        import threading
        from queue import Queue as _Queue

        self.excel_path = excel_path
        self.quota_log_path = Path(self.excel_path).with_suffix('.quota.json')
        self.lock = threading.Lock()
        self.mobile_proxy = None  # DOBAVLENO
        self.account_rotation_lock = threading.Lock()  # DOBAVLENO
        self._last_used_key = None  # DOBAVLENO
        self.cleaned_accounts = set()  # DOBAVLENO: ochishchennye akkaunty
        self._results: Dict[str, dict] = {}          # id  → result
        self._events : Dict[str, threading.Event] = {}# id  → Event dlya wait()
        limits = load_elevenlabs_limits()
        self.max_concurrent_per_account = limits.get("max_concurrent_per_account", 2)
        self.config = {
            'max_concurrent_requests': self.max_concurrent_per_account,
            'voice_id': 'EXAVITQu4vr4xnSDxMaL',  # Default voice
            'model_id': 'eleven_multilingual_v2',
            'stability': 0.5,
            'similarity_boost': 0.5,
            'style': 0.5,
            'speed': 1.0,
            'use_speaker_boost': True
        }

        # Очередь и поток обработки
        self.queue = _Queue()
        self.stop_event = threading.Event()
        self.processing_thread: threading.Thread | None = None
        self.processing = False

        # Назначения запросов на аккаунты (используется в распределении)
        self.account_assignments: Dict[str, Dict] = {}

        # Прочее
        self._ensure_excel_structure()
        self.quota_refresh_needed = False
        self.quota_refresh_accounts = set()

    
    def _ensure_excel_structure(self):
        """Sozdaet ili proveryaet strukturu Excel fayla"""
        try:
            if not os.path.exists(self.excel_path):
                log.info("📝 Creating new Excel file: %s", self.excel_path)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "ElevenLabs APIs"

                # Zagolovki
                headers = ["API Key", "Email", "Password", "Quota Remaining", "Last Checked", "Status", "Usage Count", "Total Used This Month", "Unusual Activity", "Unusual Activity Time", "Notes"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # Primer dannykh
                # Primer dannykh
                ws['A2'] = "sk_example_key_1"
                ws['B2'] = "user1@example.com"
                ws['C2'] = "password123"
                ws['D2'] = 10000
                ws['E2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ws['F2'] = "active"
                ws['G2'] = 0
                ws['H2'] = 0
                ws['I2'] = "no"
                ws['J2'] = ""
                ws['K2'] = "Zamenite na realnye dannye"

                wb.save(self.excel_path)
                log.info("✅ Excel file created")
            else:
                # Proveryaem sushchestvuyushchiy fayl
                wb = openpyxl.load_workbook(self.excel_path)
                ws = wb.active
                
                # Proveryaem nalichie neobkhodimykh stolbtsov
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = list(header_row)
                required_headers = [
                    "API Key", "Email", "Password",
                    "Quota Remaining", "Last Checked", "Status",
                    "Usage Count", "Total Used This Month",
                    "Unusual Activity", "Unusual Activity Time",
                    "Notes", "Retry Count"
                ]

                for header in required_headers:
                    if header not in headers:
                        col_index = len(headers) + 1
                        ws.cell(row=1, column=col_index, value=header)
                        headers.append(header)
                
                wb.save(self.excel_path)
                log.info("📊 Excel file loaded")
                
        except Exception as e:
            log.exception("❌ Error in _ensure_excel_structure: %s", e)

    def _load_workbook_safe(self):
        """Safely load Excel workbook, recreating if corrupted."""
        try:
            return openpyxl.load_workbook(self.excel_path)
        except BadZipFile:
            log.error(f"❌ Excel file is corrupted: {self.excel_path}. Recreating...")
            self._ensure_excel_structure()
            try:
                return openpyxl.load_workbook(self.excel_path)
            except Exception as e:
                log.error(f"❌ Failed to reload Excel file: {e}")
                raise

    def _cleanup_elevenlabs_voices_for(self, api_key: str, account_email: str):
        """Udalyaet polzovatelskie golosa s akkaunta ElevenLabs"""
        try:
            if not self.mobile_proxy:
                log.warning("⚠️ No mobile proxy available; skipping cleanup")
                return
            proxy_info = self.mobile_proxy.get_proxy_connection_info()
            if not proxy_info:
                log.warning("⚠️ Mobile proxy connection failed; skipping cleanup")
                return
            proxy_dict = {
                "http": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                "https": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
            }
            log.info(f"🌐 Using proxy for cleanup: {proxy_info['host']}:{proxy_info['port']}")

            session = requests.Session()
            session.trust_env = False
            session.proxies = proxy_dict

            headers = {
                "xi-api-key": api_key,
                "Content-Type": "application/json",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            }

            resp = session.get("https://api.elevenlabs.io/v1/voices", headers=headers, timeout=30)
            if resp.status_code != 200:
                log.error(f"❌ Voices API error: {resp.status_code} - {maybe_truncate(resp.text, 200)}")
                return

            voices = (resp.json() or {}).get("voices", [])
            voices_to_delete = []
            for v in voices:
                voice_id = v.get("voice_id")
                voice_name = v.get("name", "Unknown")
                category = v.get("category", "")
                is_owner = v.get("is_owner", False)

                if (voice_id not in BASIC_VOICE_IDS) or is_owner or (category in ["cloned", "generated", "instant_cloning"]):
                    voices_to_delete.append(v)
                    log.info(f"✅ MARKED FOR DELETION: {voice_name} (ID: {voice_id})")

            for i, v in enumerate(voices_to_delete, 1):
                vid = v.get("voice_id")
                vname = v.get("name", "Unknown")
                log.info(f"🗑️ [{i}/{len(voices_to_delete)}] Deleting: {vname} (ID: {vid})")
                try:
                    d = session.delete(f"https://api.elevenlabs.io/v1/voices/{vid}", headers=headers, timeout=30)
                    if d.status_code == 200:
                        log.info(f"✅ Deleted: {vname} (ID: {vid})")
                    elif d.status_code == 400 and "voice_does_not_exist" in (d.text or ""):
                        log.info(f"⚠️ Not found (already deleted): {vname} (ID: {vid})")
                    else:
                        log.error(f"❌ Delete failed for {vname} (ID: {vid}) — {d.status_code}: {maybe_truncate(d.text, 200)}")
                    time.sleep(1)
                except Exception as e:
                    log.error(f"❌ Exception deleting {vname} (ID: {vid}): {e}")

            try:
                session.close()
            except Exception:
                pass
        except Exception as e:
            log.error(f"❌ Cleanup error: {e}")

    def update_config(self, config):
        """Obnovlyaet konfiguratsiyu ElevenLabs menedzhera"""
        with self.lock:
            self.config.update(config)
    
    def set_mobile_proxy(self, proxy_id: str, api_key: str):
        """Ustanavlivaet mobilnyy proksi s proverkoy rabotosposobnosti"""
        try:
            temp_proxy = MobileProxyManager(proxy_id, api_key)
            
            # Proveryaem chto API rabotaet
            current_ip = temp_proxy.get_current_ip()
            if current_ip == 'unknown':
                raise Exception("Failed to get current IP - invalid credentials or API error")
            
            # Proveryaem chto mozhem poluchit informatsiyu o proksi
            stats = temp_proxy.get_stats()
            if not stats:
                raise Exception("Failed to get proxy stats - invalid proxy_id or API error")
            
            # Esli vse proverki proshli - sokhranyaem
            self.mobile_proxy = temp_proxy
            log.info(f"📱 Mobile proxy configured successfully: {proxy_id}, IP: {current_ip}")
            return True
            
        except Exception as e:
            log.error(f"❌ Failed to configure mobile proxy {proxy_id}: {e}")
            self.mobile_proxy = None
            return False

    def get_best_api_key(self, required_chars: int, avoid_unusual: bool = True, rotate_ip: bool = False) -> dict:
        """Vozvrashchaet luchshiy API klyuch s optsionalnoy filtratsiey i rotatsiey IP"""
        from contextlib import nullcontext

        key_data = None  # vazhno: initsializiruem zaranee
        outer_lock = self.account_rotation_lock if rotate_ip else nullcontext()
        with outer_lock:
            wb = None
            try:
                with self.lock:
                    wb = openpyxl.load_workbook(self.excel_path)
                    ws = wb.active

                    best_key = None
                    best_quota = -1
                    min_useful_quota = 100

                    for row in range(2, ws.max_row + 1):
                        api_key = ws[f'A{row}'].value
                        quota = ws[f'D{row}'].value
                        status = (ws[f'F{row}'].value or 'active').lower()
                        unusual_activity = (ws[f'I{row}'].value or 'no').lower()
                        email = ws[f'B{row}'].value

                        if not api_key or status == 'disabled':
                            continue
                        if avoid_unusual and unusual_activity == 'yes':
                            log.debug(f"⚠️ Skipping account with unusual activity: {email}")
                            continue

                        # esli v Excel kvota pustaya/nol — poprobuem poluchit ee cherez API odin raz
                        if quota in (None, '', 0):
                            proxy_dict = {}
                            if self.mobile_proxy:
                                proxy_info = self.mobile_proxy.get_proxy_connection_info()
                                if proxy_info:
                                    proxy_dict = {
                                        "http":  f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                                        "https": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                                    }
                            quota = self.check_quota(api_key, proxy_dict, force=True)
                            self.update_quota_in_excel(api_key, quota)

                        # otsekaem melkie ostatki
                        if quota is not None and quota != '' and quota < min_useful_quota:
                            continue

                        # esli polnostyu khvataet — berem srazu
                        if quota and quota >= required_chars:
                            key_data = {
                                'api_key': api_key,
                                'quota_remaining': quota,
                                'email': email,
                                'row': row
                            }
                            break

                        # inache — zapominaem nailuchshiy
                        if quota and quota > best_quota:
                            best_quota = quota
                            best_key = {
                                'api_key': api_key,
                                'quota_remaining': quota,
                                'email': email,
                                'row': row
                            }

                    if key_data is None:
                        key_data = best_key

            except Exception as e:
                log.error(f"❌ Error getting best API key: {e}")
                return None
            finally:
                if wb:
                    wb.close()

        if not key_data:
            log.warning("⚠️ No suitable API key found")
            return None

        if rotate_ip and self.mobile_proxy and self._last_used_key and self._last_used_key != key_data['api_key']:
            log.info(f"🔄 Account changed: {str(self._last_used_key)[-10:]} → {str(key_data['api_key'])[-10:]}")
            rotation_success = False
            for attempt in range(3):
                if self.mobile_proxy.rotate_ip():
                    log.info(f"📱 IP rotated for new account: {key_data.get('email', 'unknown')}")
                    rotation_success = True
                    break
                log.error(f"❌ IP rotation failed for new account (attempt {attempt + 1}/3)")
                time.sleep(5)
            if not rotation_success:
                log.error("❌ Unable to rotate IP after 3 attempts")
                return None

        self._last_used_key = key_data['api_key']
        return key_data


    def mark_unusual_activity(self, api_key: str, email: str = None, retry_count: int = 1):
        """Pomechaet API klyuch kak imeyushchiy podozritelnuyu aktivnost s schetchikom popytok"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                if ws[f'A{row}'].value == api_key:
                    # Poluchaem tekushchiy schetchik popytok
                    current_retries = ws[f'L{row}'].value or 0  # Novyy stolbets L dlya schetchika
                    new_retry_count = current_retries + retry_count
                    
                    ws[f'I{row}'] = "yes"  # Unusual Activity
                    ws[f'J{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Time
                    ws[f'L{row}'] = new_retry_count  # Retry count
                    
                    # Esli prevyshen limit popytok - pomechaem kak nerabochiy
                    if new_retry_count >= 4:
                        ws[f'F{row}'] = "disabled"  # Status
                        ws[f'K{row}'] = f"Disabled after {new_retry_count} unusual activity attempts"  # Notes
                        log.warning(f"🚨 Account disabled after {new_retry_count} attempts: {email}")
                    else:
                        log.warning(f"🚨 Unusual activity attempt {new_retry_count}/4: {email}")
                    
                    wb.save(self.excel_path)
                    break
                    
        except Exception as e:
            log.error(f"❌ Error marking unusual activity: {e}")

    def clear_unusual_activity(self, api_key: str):
        """Ochishchaet otmetku o podozritelnoy aktivnosti"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                if ws[f'A{row}'].value == api_key:
                    ws[f'I{row}'] = "no"
                    ws[f'J{row}'] = ""
                    wb.save(self.excel_path)
                    
                    email = ws[f'B{row}'].value or "unknown"
                    log.info(f"✅ Cleared unusual activity flag for: {email}")
                    break
                    
        except Exception as e:
            log.error(f"❌ Error clearing unusual activity: {e}")

    def mark_quota_exceeded(self, api_key: str, remaining: int = 0, message: str = ""):
        """Pomechaet akkaunt kak ischerpavshiy kvotu"""
        try:
            with self.lock:
                wb = self._load_workbook_safe()
                ws = wb.active

                for row in range(2, ws.max_row + 1):
                    if ws[f'A{row}'].value == api_key:
                        ws[f'D{row}'] = remaining
                        ws[f'E{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        if message:
                            ws[f'K{row}'] = message
                        email = ws[f'B{row}'].value or 'unknown'
                        wb.save(self.excel_path)
                        wb.close()
                        log.warning(
                            f"⚠️ Quota exceeded for {email}: remaining={remaining}"
                        )

                        # Posle oshibki proveryaem aktualnuyu kvotu i obnovlyaem Excel
                        try:
                            proxy_obj = None
                            proxy_dict = {}
                            if hasattr(g, 'proxy_manager'):
                                proxy_obj = g.proxy_manager.get_available_proxy(for_openai_fm=False)
                                proxy_dict = self._get_proxy_dict(proxy_obj) if proxy_obj else {}

                            quota_remaining = self.check_quota(api_key, proxy_dict, force=True)
                            self.update_quota_in_excel(api_key, quota_remaining)

                            # Log quota check result to JSON file
                            try:
                                data = {}
                                if self.quota_log_path.exists():
                                    with self.quota_log_path.open('r', encoding='utf-8') as f:
                                        data = json.load(f)
                                data[email] = {
                                    'api_key': api_key[-8:],
                                    'remaining': quota_remaining,
                                    'checked_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                    'message': message,
                                }
                                with self.quota_log_path.open('w', encoding='utf-8') as f:
                                    json.dump(data, f, ensure_ascii=False, indent=2)
                            except Exception as log_err:
                                log.error(f"❌ Error writing quota log: {log_err}")

                        except Exception as check_err:
                            log.error(f"❌ Error refreshing quota after exceed: {check_err}")
                        break
        except Exception as e:
            log.error(f"❌ Error marking quota exceeded: {e}")

    def check_quota(self, api_key: str, proxy_dict: dict = None, force: bool = False) -> int:
        """Proveryaet kvotu dlya API klyucha. Pri force=False prosto vozvrashchaet 0 bez zaprosa"""
        if not force:
            log.debug(f"⏭️ Skipping quota check for {api_key[-8:]}")
            return 0

        session = requests.Session()
        session.trust_env = False

        # Retrai na urovne urllib3 (na sluchay 429/5xx)
        retry = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET"]
        )
        session.mount("https://", HTTPAdapter(max_retries=retry, pool_maxsize=10))

        if proxy_dict:
            session.proxies = proxy_dict

        headers = {
            "xi-api-key": api_key,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        url = "https://api.elevenlabs.io/v1/user/subscription"

        try:
            resp = session.get(url, headers=headers)
            if resp.status_code == 200:
                data = resp.json()
                remaining = data.get('character_limit', 0) - data.get('character_count', 0)
                log.info(f"📊 API Key quota check: {remaining} characters remaining")
                return max(0, remaining)
            else:
                log.error(f"❌ Quota check failed: {resp.status_code} {maybe_truncate(resp.text, 120)}")
                return 0
        except (ReadTimeout, ConnectionError, socket.error, ConnectionResetError) as e:
            log.warning(f"⏳/RST quota check error: {e}. Will retry later.")
            return 0
        except Exception as e:
            log.error(f"❌ Error checking quota: {e}")
            return 0
        finally:
            try:
                session.close()
            except:
                pass

        
    def update_quota_in_excel(self, api_key: str, quota_remaining: int):
        """Obnovlyaet kvotu v Excel fayle"""
        try:
            with self.lock:
                wb = self._load_workbook_safe()
                ws = wb.active

                for row in range(2, ws.max_row + 1):
                    if ws[f'A{row}'].value == api_key:
                        ws[f'D{row}'] = quota_remaining
                        ws[f'E{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        break

                wb.save(self.excel_path)
                wb.close()

        except Exception as e:
            log.error(f"❌ Error updating quota in Excel: {e}")

    def refresh_all_quotas(self, accounts=None, max_workers: int = 5) -> dict:
        """Proveryaet kvotu akkauntov. Esli accounts=None - proveryayutsya vse"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            accounts_list = []
            filter_set = set(accounts) if accounts else None
            for row in range(2, ws.max_row + 1):
                api_key = ws[f'A{row}'].value
                email = ws[f'B{row}'].value
                status = ws[f'F{row}'].value or 'active'
                if not api_key or (status and status.lower() == 'disabled'):
                    continue
                if filter_set and api_key not in filter_set and email not in filter_set:
                    continue
                accounts_list.append((row, api_key, email))
            wb.close()

            results = {}

            def worker(row, api_key, email):
                proxy_dict = None
                if self.mobile_proxy:
                    proxy_info = self.mobile_proxy.get_proxy_connection_info()
                    if proxy_info:
                        proxy_dict = {
                            "http": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                            "https": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}"
                        }
                quota = self.check_quota(api_key, proxy_dict, force=True)
                with self.lock:
                    wb_local = openpyxl.load_workbook(self.excel_path)
                    ws_local = wb_local.active
                    ws_local[f'D{row}'] = quota
                    ws_local[f'E{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    wb_local.save(self.excel_path)
                    wb_local.close()
                results[email] = quota

            threads = []
            for row, api_key, email in accounts_list:
                t = threading.Thread(target=worker, args=(row, api_key, email), daemon=True)
                threads.append(t)
                t.start()
                if len(threads) >= max_workers:
                    for t in threads:
                        t.join()
                    threads = []
            for t in threads:
                t.join()
            return results
        except Exception as e:
            log.error(f"❌ Error refreshing quotas: {e}")
            return {}
        
    def _get_proxy_dict(self, proxy_obj: dict) -> dict:
            """Formiruet slovar proksi dlya requests"""
            if not proxy_obj:
                return {}
                
            cred = ""
            if proxy_obj.get("username") and proxy_obj.get("password"):
                cred = f"{proxy_obj['username']}:{proxy_obj['password']}@"
            
            proxy_url = f"http://{cred}{proxy_obj['host']}:{proxy_obj['port']}"
            return {"http": proxy_url, "https": proxy_url}
    
    def update_usage(self, api_key: str, chars_used: int):
            """Obnovlyaet ispolzovanie API klyucha"""
            try:
                wb = openpyxl.load_workbook(self.excel_path)
                ws = wb.active
                
                for row in range(2, ws.max_row + 1):
                    if ws[f'A{row}'].value == api_key:
                        current_quota = ws[f'D{row}'].value or 0
                        current_usage = ws[f'G{row}'].value or 0
                        
                        ws[f'D{row}'] = max(0, current_quota - chars_used)
                        ws[f'G{row}'] = current_usage + chars_used
                        break
                        
                wb.save(self.excel_path)
                
            except Exception as e:
                log.error(f"❌ Error updating usage: {e}")

    def check_and_update_quota_from_excel(self, api_key: str) -> int:
        """Proveryaet kvotu iz Excel; pri ustarevanii obnovlyaet cherez API"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active

            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == api_key:
                    email = ws.cell(row=row, column=2).value
                    quota_from_excel = ws.cell(row=row, column=4).value or 0
                    last_checked = ws.cell(row=row, column=5).value
                    wb.close()

                    # spolzuem kesh, esli dannye svezhie (<5 minut)
                    if quota_from_excel > 0 and last_checked:
                        try:
                            last_checked_time = datetime.strptime(last_checked, '%Y-%m-%d %H:%M:%S')
                            if (datetime.now() - last_checked_time).total_seconds() < 300:
                                log.debug(f"📊 Using cached quota for {email}: {quota_from_excel}")
                                return quota_from_excel
                        except Exception:
                            pass

                    # Kvota otsutstvuet ili ustarela – obnovlyaem cherez API
                    new_quota = self.check_quota(api_key, force=True)
                    self.update_quota_in_excel(api_key, new_quota)
                    log.debug(f"📊 Refreshed quota for {email}: {new_quota}")
                    return new_quota

            wb.close()
            return 0

        except Exception as e:
            log.error(f"❌ Error checking quota from Excel: {e}")
            return 0

    def _calculate_quota_cost(self, chars_used: int, model_id: str) -> int:
        """Vozvrashchaet stoimost zaprosa v simvolakh s uchetom modeli"""
        if model_id in DISCOUNTED_MODELS:
            return math.ceil(chars_used / 2)
        return chars_used

    def update_quota_after_request(self, api_key: str, chars_used: int, model_id: str):
        """Obnovlyaet kvotu v Excel posle otpravki zaprosa (potokobezopasno)"""
        cost = self._calculate_quota_cost(chars_used, model_id)
        with self.lock:
            try:
                wb = openpyxl.load_workbook(self.excel_path)
                ws = wb.active

                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == api_key:
                        email = ws.cell(row=row, column=2).value
                        current_quota = ws.cell(row=row, column=4).value or 0
                        current_usage = ws.cell(row=row, column=7).value or 0

                        new_quota = max(0, current_quota - cost)
                        new_usage = current_usage + cost

                        ws.cell(row=row, column=4, value=new_quota)
                        ws.cell(row=row, column=7, value=new_usage)
                        ws.cell(row=row, column=5, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

                        wb.save(self.excel_path)
                        wb.close()

                        log.debug(f"📊 Updated quota for {email}: {new_quota} remaining (+{cost} used)")
                        break

            except Exception as e:
                log.error(f"❌ Error updating quota after request: {e}")



    def ensure_account_voices_cleaned(self, api_key: str, account_email: str = None, proxies: dict | None = None):
        """Ensure cleanup of custom/generated voices for this API key.
        
        This routine previously ran only once per process which allowed
        accounts to accumulate voices again and eventually hit the
        ``voice_limit_reached`` error.  Now it executes on every call but
        is lightly throttled to avoid duplicate cleanup when invoked
        multiple times in quick succession.
        """
        try:
            # Track last cleanup time per API key so rapid consecutive calls
            # (e.g. retry logic) do not hammer the ElevenLabs API.
            if not hasattr(self, "_last_voice_cleanup"):
                self._last_voice_cleanup = {}
            last_time = self._last_voice_cleanup.get(api_key)
            if last_time and time.time() - last_time < 10:
                return

            session = requests.Session()
            session.trust_env = False
            # Prefer explicit proxies; else mobile proxy; else no proxies
            if proxies:
                session.proxies = proxies
            elif self.mobile_proxy:
                proxy_info = self.mobile_proxy.get_proxy_connection_info()
                if proxy_info:
                    session.proxies = {
                        "http": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                        "https": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}"
                    }

            headers = {
                "xi-api-key": api_key,
                "Content-Type": "application/json",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            }

            resp = session.get("https://api.elevenlabs.io/v1/voices", headers=headers, timeout=30)
            if resp.status_code != 200:
                log.error(f"Voices API error (cleanup): {resp.status_code} - {maybe_truncate(resp.text, 200)}")
                try:
                    session.close()
                except Exception:
                    pass
                return

            voices = (resp.json() or {}).get("voices", [])
            deletable = []
            for v in voices:
                vid = v.get("voice_id")
                category = v.get("category", "")
                is_owner = v.get("is_owner", False)
                if (vid not in BASIC_VOICE_IDS) or is_owner or (category in ["cloned", "generated", "instant_cloning"]):
                    deletable.append(v)

            for v in deletable:
                vid = v.get("voice_id")
                name = v.get("name", "Unknown")
                try:
                    d = session.delete(f"https://api.elevenlabs.io/v1/voices/{vid}", headers=headers, timeout=30)
                    if d.status_code == 200:
                        log.info(f"Deleted custom voice: {name} ({vid})")
                    elif d.status_code == 400 and "voice_does_not_exist" in (d.text or ""):
                        log.info(f"Voice already deleted: {name} ({vid})")
                    else:
                        log.warning(f"Voice delete failed {name} ({vid}): {d.status_code} {maybe_truncate(d.text, 120)}")
                    time.sleep(0.5)
                except Exception as de:
                    log.warning(f"Voice delete exception {name} ({vid}): {de}")

            try:
                session.close()
            except Exception:
                pass

            # Record cleanup time for this API key
            self._last_voice_cleanup[api_key] = time.time()
        except Exception as e:
            log.error(f"ensure_account_voices_cleaned error: {e}")


class ElevenLabsQueue:
    def __init__(self, excel_path: str = "api_elevenlabs.xlsx"):
        self.excel_path = excel_path
        self.queue = thread_queue.Queue()
        self.processing = False
        self.lock = threading.Lock()
        self.account_assignments = {}  # {account_id: [requests]}
        self.processing_thread = None
        self._events: Dict[str, threading.Event] = {}
        self._results: Dict[str, dict] = {}
        self.stop_event = threading.Event()

        # Default configuration for requests in the queue. This mirrors
        # the defaults used by ``ElevenLabsManager`` so that queued requests
        # have sensible parameters even if no explicit config is supplied.
        self.config = {
            "model_id": "eleven_multilingual_v2",
            **VOICE_DEFAULTS,
        }

        self.quota_refresh_needed = False
        self.quota_refresh_accounts = set()

        # Mobile proxy instance reused across managers (set externally)
        self.mobile_proxy = None
        limits = load_elevenlabs_limits()
        self._account_semaphores = {}  # api_key -> Semaphore(limit)
        self._semaphores_lock = threading.Lock()
        self._max_concurrent_per_account = limits.get("max_concurrent_per_account", 2)
        self._batch_size = limits.get("batch_size")
        self._cleanup_events_lock = threading.Lock()
        self._account_cleanup_events = {}


    def update_config(self, config: dict) -> None:
        """Update the queue's default configuration."""
        if not isinstance(config, dict):
            return

        # Обновляем общий конфиг
        self.config.update(config)

    def _get_account_semaphore(self, api_key: str) -> threading.Semaphore:
        """Return per-account semaphore, creating it with configured limit."""
        with self._semaphores_lock:
            sem = self._account_semaphores.get(api_key)
            if sem is None:
                sem = threading.Semaphore(self._max_concurrent_per_account)
                self._account_semaphores[api_key] = sem
            return sem



    def update_quota_after_request(self, api_key: str, chars_used: int, model_id: str):
        """Delegate quota update to the global manager if available."""
        manager = getattr(g, "elevenlabs_manager", None)
        if manager and hasattr(manager, "update_quota_after_request"):
            manager.update_quota_after_request(api_key, chars_used, model_id)

    def mark_quota_exceeded(self, api_key: str, remaining: int = 0, message: str = ""):
        """Delegate quota exceeded marking to the global manager."""
        manager = getattr(g, "elevenlabs_manager", None)
        if manager and hasattr(manager, "mark_quota_exceeded"):
            manager.mark_quota_exceeded(api_key, remaining, message)

    def _ensure_initial_voice_cleanup(self, account: dict, proxies: dict | None):
        """Perform voice cleanup once per account for the current batch.

        Subsequent requests for the same account wait until cleanup finishes.
        """
        api_key = account.get("api_key")
        if not api_key:
            return
        with self._cleanup_events_lock:
            event = self._account_cleanup_events.get(api_key)
            if event is None:
                event = threading.Event()
                self._account_cleanup_events[api_key] = event
                run_cleanup = True
            else:
                run_cleanup = False
        if run_cleanup:
            try:
                self.ensure_account_voices_cleaned(api_key, account.get("email"), proxies)
            finally:
                event.set()
        else:
            event.wait()

    def ensure_account_voices_cleaned(self, api_key: str, account_email: str = None, proxies: dict | None = None):
        """Remove custom/generated voices for this API key (throttled)."""
        try:
            if not hasattr(self, "_last_voice_cleanup"):
                self._last_voice_cleanup = {}
            last_time = self._last_voice_cleanup.get(api_key)
            if last_time and time.time() - last_time < 10:
                return

            session = requests.Session()
            session.trust_env = False
            if proxies:
                session.proxies = proxies
            elif self.mobile_proxy:
                proxy_info = self.mobile_proxy.get_proxy_connection_info()
                if proxy_info:
                    session.proxies = {
                        "http": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                        "https": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                    }

            headers = {
                "xi-api-key": api_key,
                "Content-Type": "application/json",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            }

            resp = session.get("https://api.elevenlabs.io/v1/voices", headers=headers, timeout=30)
            if resp.status_code != 200:
                log.error(f"Voices API error (cleanup): {resp.status_code} - {maybe_truncate(resp.text, 200)}")
                session.close()
                return

            voices = (resp.json() or {}).get("voices", [])
            deletable = []
            for v in voices:
                vid = v.get("voice_id")
                category = v.get("category", "")
                is_owner = v.get("is_owner", False)
                if (vid not in BASIC_VOICE_IDS) or is_owner or (category in ["cloned", "generated", "instant_cloning"]):
                    deletable.append(v)

            for v in deletable:
                vid = v.get("voice_id")
                name = v.get("name", "Unknown")
                try:
                    d = session.delete(f"https://api.elevenlabs.io/v1/voices/{vid}", headers=headers, timeout=30)
                    if d.status_code == 200:
                        log.info(f"Deleted custom voice: {name} ({vid})")
                    elif d.status_code == 400 and "voice_does_not_exist" in (d.text or ""):
                        log.info(f"Voice already deleted: {name} ({vid})")
                    else:
                        log.warning(f"Voice delete failed {name} ({vid}): {d.status_code} {maybe_truncate(d.text, 120)}")
                    time.sleep(0.5)
                except Exception as de:
                    log.warning(f"Voice delete exception {name} ({vid}): {de}")

            session.close()
            self._last_voice_cleanup[api_key] = time.time()
        except Exception as e:
            log.error(f"ensure_account_voices_cleaned error: {e}")


    def add_request(self, text: str, voice_id: str, config: dict) -> str:
        """Kladet zapros v ochered i sozdaet Event dlya ozhidaniya."""
        req_id = str(uuid.uuid4())
        chars_needed = len(text)

        if self.stop_event.is_set():
            log.warning("⚠️ Queue stopped, rejecting request %s", req_id[:8])
            self._events[req_id] = threading.Event()
            result = {'success': False, 'error': 'Queue stopped'}
            self._store_result(req_id, result)
            return req_id

        req = {
            "id": req_id,
            "text": text,
            "voice_id": voice_id,
            "config": config,
            "chars_needed": chars_needed,
            "status": "queued",
            "result": None,
        }
        self._events[req_id] = threading.Event()
        self.queue.put(req)
        log.info("📝 Request %s queued: %d chars", req_id[:8], chars_needed)
        self._start_processing()
        return req_id

    def _calculate_quota_cost(self, chars_used: int, model_id: str) -> int:
        """Return the quota cost for a request taking model discounts into account."""
        if model_id in DISCOUNTED_MODELS:
            return math.ceil(chars_used / 2)
        return chars_used

    def _store_result(self, request_id: str, result: dict) -> None:
        """Kladet rezultat i budit ozhidayushchiy potok."""
        self._results[request_id] = result
        ev = self._events.get(request_id)
        if ev:
            ev.set()

    def wait_for_result(self, request_id: str, timeout: int = 300) -> dict | None:
        """Blokiruetsya do rezultata ili taym-auta."""
        ev = self._events.get(request_id)
        if not ev:
            return None
        is_set = ev.wait(timeout)
        return self._results.get(request_id) if is_set else None


    def _get_additional_accounts(self, needed_quota: int, max_single_needed: int = 0):
        """shchet dopolnitelnye akkaunty s pustoy kvotoy dlya proverki"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            additional_accounts = []
            found_quota = 0
            checked_count = 0
            max_check = 20  # Maksimum 20 dopolnitelnykh akkauntov za raz
            found_suitable_single = False
            
            log.info(f"🔍 Looking for additional accounts (need {needed_quota} more quota, max_single={max_single_needed})")
            
            for row in range(2, ws.max_row + 1):
                if checked_count >= max_check:
                    log.info(f"⏸️ Reached max check limit ({max_check}), stopping additional search")
                    break
                    
                # SPRAVLENE: Proveryaem usloviya ostanovki na kazhdoy iteratsii
                if found_quota >= needed_quota and (max_single_needed == 0 or found_suitable_single):
                    log.info(f"✅ Found enough quota during search: {found_quota} >= {needed_quota}, suitable_single={found_suitable_single}")
                    break
                    
                api_key = ws[f'A{row}'].value
                email = ws[f'B{row}'].value
                status = ws[f'F{row}'].value or 'active'
                unusual_activity = ws[f'I{row}'].value or 'no'
                quota = ws[f'D{row}'].value

                if not api_key or (status and status.lower() == 'disabled') or unusual_activity == 'yes':
                    continue
                
                # shchem tolko akkaunty s pustoy kvotoy (kotorye eshche ne proveryalis)
                if quota is None or quota == '' or quota == 0:
                    log.info(f"🔍 Checking additional account: {email}")
                    checked_count += 1
                    
                    if self.mobile_proxy:
                        proxy_info = self.mobile_proxy.get_proxy_connection_info()
                        if proxy_info:
                            proxy_dict = {
                                "http": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                                "https": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}"
                            }
                        else:
                            continue
                    else:
                        continue

                    new_quota = self.check_quota(api_key, proxy_dict, force=True)
                    
                    # Sokhranyaem v Excel
                    ws[f'D{row}'] = new_quota
                    ws[f'E{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    wb.save(self.excel_path)
                    
                    log.info(f"💾 Additional account {email}: {new_quota} quota")
                    
                    if new_quota and new_quota > 0:
                        additional_accounts.append({
                            'api_key': api_key,
                            'email': email,
                            'quota_remaining': new_quota,
                            'row': row
                        })
                        found_quota += new_quota
                        
                        # Proveryaem podkhodit li dlya bolshogo zaprosa
                        if max_single_needed > 0 and new_quota >= max_single_needed:
                            found_suitable_single = True
                            log.info(f"✅ Found account suitable for large request: {email} ({new_quota} >= {max_single_needed})")
                    
                    time.sleep(2)  # Pauza mezhdu zaprosami
            
            wb.close()
            
            # Sortiruem po ubyvaniyu kvoty
            additional_accounts.sort(key=lambda x: x['quota_remaining'], reverse=True)
            
            log.info(f"📊 Additional search result: found {len(additional_accounts)} accounts with {found_quota} total quota")
            
            return additional_accounts
            
        except Exception as e:
            log.error(f"❌ Error getting additional accounts: {e}")
            return []
        
# elevenlabs_manager.py  ───────────────────────────────────────────────
    def _check_quota_requirements(self, requests) -> tuple[bool, list]:
        """
        Vozvrashchaet (True, accounts), esli:
        • summarnaya dostupnaya kvota ≥ summarnoy potrebnosti;
        • sushchestvuet akkaunt, chya kvota ≥ samomu bolshomu zaprosu.
        nache pytaetsya nayti dopolnitelnye akkaunty i pereschityvaet.
        """
        total_chars_needed  = sum(r['chars_needed'] for r in requests)
        max_single_request  = max((r['chars_needed'] for r in requests), default=0)

        log.info(f"📊 Quota requirements: total={total_chars_needed}, "
                f"max_single={max_single_request}")

        # 1️⃣ Berem dostupnye (bystryy prokhod s rannim break)
        accounts = self._get_available_accounts(required_quota=total_chars_needed)
        total_available   = sum(a['quota_remaining'] for a in accounts)
        max_account_quota = max((a['quota_remaining'] for a in accounts), default=0)

        has_enough_total   = total_available   >= total_chars_needed
        has_big_enough_one = max_account_quota >= max_single_request
        log.info(f"📊 Found {len(accounts)} accounts: total={total_available}, "
                f"max_account={max_account_quota}")

        # 2️⃣ Esli summarno khvataet, no net «tolstogo» akkaunta – doskaniruem fayl
        if has_enough_total and not has_big_enough_one:
            log.warning("⚠️ Total quota OK, no ni odin akkaunt ne pomeshchaet "
                        "krupnyy zapros – prodolzhayu skanirovanie Excel")
            full_scan = self._get_available_accounts(force_refresh=True, required_quota=0)
            # dobavlyaem novye, izbegaya dublikatov
            known = {a['api_key']: a for a in accounts}
            for acc in full_scan:
                known.setdefault(acc['api_key'], acc)
            accounts = list(known.values())
            max_account_quota = max(a['quota_remaining'] for a in accounts)
            has_big_enough_one = max_account_quota >= max_single_request
            log.info(f"📊 Posle polnogo skanirovaniya max_account={max_account_quota}")

        # 3️⃣ Esli po-prezhnemu net nuzhnogo akkaunta – proveryaem ne-obnovlennye
        if (not has_enough_total) or (not has_big_enough_one):
            need_more = max(0, total_chars_needed - total_available)
            extra = self._get_additional_accounts(need_more, 
                                                0 if has_big_enough_one else max_single_request)
            if extra:
                known = {a['api_key']: a for a in accounts}
                for acc in extra:
                    known.setdefault(acc['api_key'], acc)
                accounts = list(known.values())
                total_available   = sum(a['quota_remaining'] for a in accounts)
                max_account_quota = max(a['quota_remaining'] for a in accounts)
                has_enough_total   = total_available   >= total_chars_needed
                has_big_enough_one = max_account_quota >= max_single_request
                log.info(f"📊 After extra search: total={total_available}, "
                        f"max_account={max_account_quota}")

        if has_enough_total and has_big_enough_one:
            log.info("✅ Quota requirements satisfied (total & single)")
            return True, accounts

        log.error(f"❌ Insufficient quota — need_total={total_chars_needed}, "
                f"have_total={total_available}; need_single={max_single_request}, "
                f"have_max_account={max_account_quota}")
        return False, accounts
# ──────────────────────────────────────────────────────────────────────


    def _start_processing(self):
        """Zapuskaet obrabotku ocheredi esli eshche ne zapushchena"""
        with self.lock:
            if not self.processing:
                self.stop_event.clear()
                self.processing = True
                self.processing_thread = threading.Thread(target=self._process_queue, daemon=True)
                self.processing_thread.start()
                log.info("🚀 Started ElevenLabs queue processing")

    def _process_queue(self):
        """Osnovnoy tsikl obrabotki ocheredi"""
        while not self.stop_event.is_set():
            try:
                # Zhdem khotya by odin zapros v ocheredi
                try:
                    first_req = self.queue.get(timeout=2)
                except thread_queue.Empty:
                    continue

                requests = [first_req]

                # Sbor dopolnitelnykh zaprosov s ogranicheniem batch_size
                while self._batch_size is None or len(requests) < self._batch_size:
                    try:
                        req = self.queue.get(timeout=2)
                        requests.append(req)
                    except thread_queue.Empty:
                        break

                log.info(f"📋 Processing batch of {len(requests)} requests")

                # Zaprashivaem informatsiyu o proksi odin raz dlya vsey pachki
                if self.mobile_proxy:
                    stats = self.mobile_proxy.get_stats()
                    self.mobile_proxy.get_proxy_connection_info(stats)
                
                # DOBAVT: Proveryaem khvataet li obshchey kvoty
                quota_ok, available_accounts = self._check_quota_requirements(requests)
                if not quota_ok:
                    log.error("❌ Cannot process requests - insufficient quota across all accounts")
                    # Mozhno pometit zaprosy kak failed ili podozhdat
                    time.sleep(60)  # Zhdem minutu i probuem snova
                    continue

                # Raspredelyaem zaprosy po akkauntam
                self._assign_requests_to_accounts(requests, available_accounts)

                # Reset per-batch voice cleanup trackers
                self._account_cleanup_events = {}

                # Obrabatyvaem akkaunty po ocheredi
                if self.stop_event.is_set():
                    break

                suspicious = self._process_accounts_concurrently()

                if suspicious:
                    log.warning("🚨 Suspicious activity detected in batch; rotating proxy")
                    if self.mobile_proxy:
                        try:
                            self.mobile_proxy.rotate_ip()
                        except Exception as exc:
                            log.error(f"⚠️ Proxy rotation failed: {exc}")
                    # Сбрасываем и возвращаем подозрительные в очередь
                    for req in suspicious:
                        req["status"] = "pending"
                        req["result"] = None
                        self.queue.put(req)


            except Exception as e:
                log.error(f"❌ Queue processing error: {e}")
                time.sleep(5)

    def stop(self):
        """Останавливает обработку очереди немедленно"""
        with self.lock:
            self.stop_event.set()
            self.processing = False
            try:
                while not self.queue.empty():
                    self.queue.get_nowait()
            except Exception:
                pass

        if getattr(self, "processing_thread", None) and self.processing_thread.is_alive():
            self.processing_thread.join(timeout=5)

        log.info("🛑 ElevenLabs queue processing stopped")

    def _assign_requests_to_accounts(self, requests, available_accounts):
        """Raspredelyaet zaprosy po akkauntam na osnove dostupnoy kvoty"""
        log.info("📊 Assigning requests to accounts...")

        self.account_assignments = {}
        unassigned_requests = sorted(requests, key=lambda x: x['chars_needed'])
        log.info(f"📋 Sorted requests by size: {[req['chars_needed'] for req in unassigned_requests]}")

        # Pytaemsya naznachit snachala samye malenkie zaprosy
        for req in unassigned_requests[:]:
            req_chars = req['chars_needed']
            model_id = req.get('config', {}).get('model_id', self.config.get('model_id'))
            cost = self._calculate_quota_cost(req_chars, model_id)
            assigned = False

            # shchem podkhodyashchiy akkaunt sverkhu vniz po spisku
            for account in available_accounts:
                account_id = account['api_key']
                available_quota = account['quota_remaining']

                used_quota = 0
                if account_id in self.account_assignments:
                    used_quota = self.account_assignments[account_id]['total_chars']

                remaining_quota = available_quota - used_quota

                if cost <= remaining_quota:
                    if account_id not in self.account_assignments:
                        self.account_assignments[account_id] = {
                            'account': account,
                            'requests': [],
                            'total_chars': 0
                        }

                    self.account_assignments[account_id]['requests'].append(req)
                    self.account_assignments[account_id]['total_chars'] += cost
                    unassigned_requests.remove(req)
                    assigned = True
                    log.info(
                        f"✅ Assigned request {req['id'][:8]} ({req_chars} chars) to {account['email']} "
                        f"(remaining: {remaining_quota - cost})"
                    )
                    break

            if not assigned:
                log.warning(
                    f"⚠️ Could not assign request {req['id'][:8]} ({req_chars} chars) - no account with sufficient quota"
                )
        
        # Logiruem itogovye naznacheniya
        for account_id, assignment in self.account_assignments.items():
            account = assignment['account']
            requests_count = len(assignment['requests'])
            used_quota = assignment['total_chars']
            available_quota = account['quota_remaining']

            log.info(f"📋 Account {account['email']}: {requests_count} requests, {used_quota}/{available_quota} chars used")
        
        # Esli ostalis nenaznachennye zaprosy - pokazyvaem pochemu
        if unassigned_requests:
            log.error(f"❌ {len(unassigned_requests)} requests could not be assigned:")
            for req in unassigned_requests:
                log.error(f"  📄 Request {req['id'][:8]}: {req['chars_needed']} chars needed")
            
            # Pokazyvaem dostupnye kvoty dlya diagnostiki
            log.info("📊 Available account quotas for unassigned requests:")
            for account in available_accounts:
                account_id = account['api_key']
                used_quota = 0
                if account_id in self.account_assignments:
                    used_quota = self.account_assignments[account_id]['total_chars']
                remaining = account['quota_remaining'] - used_quota
                log.info(f"  💰 {account['email']}: {remaining} chars remaining (total: {account['quota_remaining']})")

    def _get_available_accounts(self, force_refresh: bool = False, required_quota: int = 0):
        """Poluchaet spisok dostupnykh akkauntov s proverkoy aktualnoy kvoty"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active

            accounts = []
            processed_count = 0
            total_quota_collected = 0
            max_account_quota = 0
            changes_made = False

            skipped_status = []
            skipped_low_quota = []
            cached_quota_accounts = []
            refreshed_quota_accounts = []
            available_accounts_info = []

            min_useful_quota = 100
            log.info(f"📊 Scanning Excel file for accounts (required_quota={required_quota}, min_useful={min_useful_quota})")

            for row in range(2, ws.max_row + 1):
                api_key = ws[f'A{row}'].value
                email = ws[f'B{row}'].value
                status = ws[f'F{row}'].value or 'active'
                unusual_activity = ws[f'I{row}'].value or 'no'

                if not api_key:
                    continue

                processed_count += 1

                if (status and status.lower() == 'disabled') or unusual_activity == 'yes':
                    skipped_status.append(f"{email} (status={status}, unusual={unusual_activity})")
                    continue

                last_checked = ws[f'E{row}'].value
                quota = ws[f'D{row}'].value

                if quota is not None and quota != '' and quota < min_useful_quota:
                    skipped_low_quota.append(f"{email} ({quota})")
                    continue

                if quota and quota >= min_useful_quota:# and not self._is_quota_check_needed(last_checked):
                    cached_quota_accounts.append(f"{email} ({quota})")
                else:
                    should_check_quota = force_refresh or quota is None or quota == 0 or quota == ''
                    if should_check_quota:
                        if self.mobile_proxy:
                            proxy_info = self.mobile_proxy.get_proxy_connection_info()
                            if proxy_info:
                                proxy_dict = {
                                    "http": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                                    "https": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}"
                                }
                            else:
                                log.error("❌ No proxy connection info for quota check")
                                continue
                        else:
                            log.error("❌ No mobile proxy available for quota check")
                            continue

                        quota = self.check_quota(api_key, proxy_dict, force=True)

                        ws[f'D{row}'] = quota
                        ws[f'E{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        changes_made = True
                        wb.save(self.excel_path)
                        refreshed_quota_accounts.append(f"{email} ({quota})")
                        time.sleep(2)

                if quota and quota >= min_useful_quota:
                    accounts.append({
                        'api_key': api_key,
                        'email': email,
                        'quota_remaining': quota,
                        'row': row
                    })
                    total_quota_collected += quota
                    max_account_quota = max(max_account_quota, quota)
                    available_accounts_info.append(f"{email} ({quota})")
                else:
                    skipped_low_quota.append(f"{email} ({quota})")

            if changes_made:
                wb.save(self.excel_path)

            wb.close()

            # Wait a moment to group logs after burst
            time.sleep(1)

            # if skipped_status:
            #     log.debug(f"⏭️ Skipped accounts (status/unusual): {', '.join(skipped_status)}")
            # if skipped_low_quota:
            #     log.debug(f"⏭️ Skipped accounts with low quota: {', '.join(skipped_low_quota)}")
            # if cached_quota_accounts:
            #     log.debug(f"✅ Using cached quota for: {', '.join(cached_quota_accounts)}")
            if refreshed_quota_accounts:
                log.info(f"💾 Updated quota for: {', '.join(refreshed_quota_accounts)}")
            # if available_accounts_info:
                # log.info(f"✅ Accounts with available quota: {', '.join(available_accounts_info)}")

            accounts.sort(key=lambda x: x['row'])

            log.info(f"📊 Search complete: processed {processed_count} rows, found {len(accounts)} accounts")
            log.info(f"📊 Total quota: {total_quota_collected}, max single: {max_account_quota}")

            return accounts

        except Exception as e:
            log.error(f"❌ Error getting available accounts: {e}")
            return []

    def _reassign_overflow_requests(self, unassigned_requests, available_accounts):
        """Perenaznachaet nenaznachennye zaprosy na akkaunty s dostatochnoy kvotoy"""
        log.info(f"🔄 Trying to reassign {len(unassigned_requests)} overflow requests")
        
        for req in unassigned_requests[:]:
            req_chars = req['chars_needed']
            log.info(f"🔍 Looking for account for request {req['id'][:8]} ({req_chars} chars)")
            
            # shchem akkaunt s dostatochnoy SVOBODNOY kvotoy
            best_account = None
            min_remaining_quota = float('inf')
            
            for account in available_accounts:
                account_id = account['api_key']
                account_quota = account['quota_remaining']
                
                # Vychislyaem uzhe ispolzovannuyu kvotu etogo akkaunta
                used_quota = 0
                if account_id in self.account_assignments:
                    used_quota = self.account_assignments[account_id]['total_chars']
                
                remaining_quota = account_quota - used_quota
                
                log.debug(f"  🔍 {account['email']}: total={account_quota}, used={used_quota}, remaining={remaining_quota}")
                
                # Proveryaem pomestitsya li zapros v ostavshuyusya kvotu
                if remaining_quota >= req_chars:
                    # Vybiraem akkaunt s naimenshey ostavsheysya kvotoy (chtoby effektivnee ispolzovat resursy)
                    if remaining_quota < min_remaining_quota:
                        min_remaining_quota = remaining_quota
                        best_account = account
            
            if best_account:
                account_id = best_account['api_key']
                
                # Dobavlyaem zapros k sushchestvuyushchemu naznacheniyu ili sozdaem novoe
                if account_id not in self.account_assignments:
                    self.account_assignments[account_id] = {
                        'account': best_account,
                        'requests': [],
                        'total_chars': 0
                    }
                
                self.account_assignments[account_id]['requests'].append(req)
                self.account_assignments[account_id]['total_chars'] += req_chars
                unassigned_requests.remove(req)
                
                log.info(f"📋 Overflow: assigned request {req['id'][:8]} ({req_chars} chars) to {best_account['email']}")
            else:
                log.error(f"❌ No account found for request {req['id'][:8]} ({req_chars} chars)")

    def _process_all_requests_for_account(self, account: dict, requests: list):
        """Obrabatyvaet vse zaprosy dlya akkaunta posle uspeshnogo testa"""
        log.info(f"🎵 Processing {len(requests)} requests for {account['email']}")
        
        for req in requests:
            if req['status'] == 'completed':
                log.info(f"✅ Request {req['id']} already completed during test")
                continue  # Pervyy zapros uzhe obrabotan
                
            log.info(f"🎵 Processing request {req['id']} for {account['email']}")
            result = self._make_elevenlabs_request(account, req)
            req['result'] = result
            req['status'] = 'completed' if result['success'] else 'failed'
            self._store_result(req['id'], result)

            if result['success']:
                log.info(f"✅ Request {req['id']} completed successfully")
            else:
                log.error(f"❌ Request {req['id']} failed: {result.get('error', 'Unknown error')}")

            # Nebolshaya pauza mezhdu zaprosami k odnomu akkauntu
            time.sleep(1)
        
        log.info(f"🎵 Completed all {len(requests)} requests for {account['email']}")

    def _get_already_checked_accounts_for_reassignment(self, failed_account_email: str, required_quota: int):
        """Poluchaet uzhe proverennye akkaunty dlya perenaznacheniya (bez dopolnitelnykh API zaprosov)"""
        try:
            with self.lock:
                wb = self._load_workbook_safe()
                ws = wb.active

            accounts = []
            failed_api_key = None
            
            # Nakhodim API klyuch problemnogo akkaunta
            for row in range(2, ws.max_row + 1):
                email = ws[f'B{row}'].value
                if email == failed_account_email:
                    failed_api_key = ws[f'A{row}'].value
                    break
            
            # Sobiraem podkhodyashchie akkaunty
            for row in range(2, ws.max_row + 1):
                api_key = ws[f'A{row}'].value
                email = ws[f'B{row}'].value
                status = ws[f'F{row}'].value or 'active'
                unusual_activity = ws[f'I{row}'].value or 'no'
                quota = ws[f'D{row}'].value
                last_checked = ws[f'E{row}'].value

                if not api_key or api_key == failed_api_key:
                    continue
                    
                if (status and status.lower() == 'disabled') or unusual_activity == 'yes':
                    log.debug(f"⏭️ Skipping {email}: status={status}, unusual={unusual_activity}")
                    continue
                
                # Proveryaem chto kvota svezhaya i dostatochnaya
                if quota and quota > 0: #and not self._is_quota_check_needed(last_checked):  # SPRAVLENE: ispolzuem self
                    if quota >= required_quota:
                        log.info(f"✅ Found cached account with sufficient quota: {email} ({quota} >= {required_quota})")
                        accounts.append({
                            'api_key': api_key,
                            'email': email,
                            'quota_remaining': quota,
                            'row': row
                        })
                        break  # Nashli odin podkhodyashchiy - dostatochno
            
            wb.close()
            return accounts
            
        except Exception as e:
            log.error(f"❌ Error getting cached accounts for reassignment: {e}")
            return []

    def _get_api_key_by_email(self, email: str):
        """Poluchaet API klyuch po email"""
        try:
            with self.lock:
                wb = self._load_workbook_safe()
                ws = wb.active

                for row in range(2, ws.max_row + 1):
                    if ws[f'B{row}'].value == email:
                        api_key = ws[f'A{row}'].value
                        wb.close()
                        return api_key

                wb.close()
            return None

        except Exception as e:
            log.error(f"❌ Error getting API key by email: {e}")
            return None

    def _reassign_failed_requests(self, failed_requests: list, failed_account_email: str):
        """Perenaznachaet zaprosy s problemnogo akkaunta. BEZ zapuska otdelnogo potoka —
        novye naznacheniya popadut v osnovnoy tsikl obrabotki.
        """
        log.warning(f"🔄 Reassigning {len(failed_requests)} requests from failed account {failed_account_email}")

        total_chars_needed = sum(len(req['text']) for req in failed_requests)
        log.info(f"📊 Need {total_chars_needed} characters for reassignment")

        available_accounts = self._get_already_checked_accounts_for_reassignment(
            failed_account_email, total_chars_needed
        )
        if not available_accounts:
            log.info("📊 No suitable accounts found in cache, checking more accounts...")
            available_accounts = self._get_all_available_accounts_for_reassignment(
                required_quota=total_chars_needed
            )
            failed_api_key = self._get_api_key_by_email(failed_account_email)
            if failed_api_key:
                available_accounts = [acc for acc in available_accounts if acc['api_key'] != failed_api_key]
                log.info(f"📊 Excluded failed account {failed_account_email} from reassignment")

        if not available_accounts:
            log.error("❌ No accounts available for reassignment after quota check")
            for req in failed_requests:
                req['status'] = 'failed'
                req['result'] = {'success': False, 'error': 'No available accounts'}
                self._store_result(req['id'], {'success': False, 'error': 'No available accounts'})
            return

        log.info(f"📊 Found {len(available_accounts)} accounts for reassignment")
        new_assignments = {}

        def add_assignment(acc, reqs, total_chars):
            api_key = acc['api_key']
            if api_key in self.account_assignments:
                self.account_assignments[api_key]['requests'].extend(reqs)
                self.account_assignments[api_key]['total_chars'] += total_chars
            elif api_key in new_assignments:
                new_assignments[api_key]['requests'].extend(reqs)
                new_assignments[api_key]['total_chars'] += total_chars
            else:
                new_assignments[api_key] = {
                    'account': acc,
                    'requests': reqs.copy(),
                    'total_chars': total_chars
                }

        # Snachala — popytka odnim akkauntom
        single_ok = False
        for account in available_accounts:
            if account['quota_remaining'] >= total_chars_needed:
                add_assignment(account, failed_requests, total_chars_needed)
                failed_requests.clear()
                single_ok = True
                log.info(
                    f"✅ Found single account with sufficient quota: {account['email']} "
                    f"({account['quota_remaining']} >= {total_chars_needed})"
                )
                break

        # nache — raspredelyaem po neskolkim
        if not single_ok:
            remaining = failed_requests.copy()
            for account in available_accounts:
                if not remaining:
                    break
                cap = account['quota_remaining']
                pack, used = [], 0
                for req in remaining[:]:
                    need = len(req['text'])
                    if used + need <= cap:
                        pack.append(req)
                        used += need
                        remaining.remove(req)
                if pack:
                    add_assignment(account, pack, used)
                    log.info(f"📋 Assigned {len(pack)} requests to {account['email']} ({used} chars)")
            failed_requests[:] = remaining

        if failed_requests:
            log.error(
                f"❌ {len(failed_requests)} requests could not be reassigned - insufficient total quota"
            )
            for req in failed_requests:
                req['status'] = 'failed'
                req['result'] = {
                    'success': False,
                    'error': 'Insufficient total quota across all accounts'
                }
                self._store_result(
                    req['id'],
                    {'success': False, 'error': 'Insufficient total quota across all accounts'}
                )

        if new_assignments:
            log.info(
                f"🔄 Merging {len(new_assignments)} reassigned account assignments into main queue"
            )
            # Prosto dobavlyaem k osnovnym naznacheniyam — oni budut obrabotany na sleduyushchem prokhode
            requeued = 0
            for data in new_assignments.values():
                for r in data.get('requests', []):
                    r['status'] = 'pending'
                    r['result'] = None
                    self.queue.put(r)
                    requeued += 1
            log.info(f"Requeued {requeued} requests for next batch after reassignment")


    def _get_all_available_accounts_for_reassignment(self, required_quota: int = 0):
        """Poluchaet dostupnye akkaunty dlya perenaznacheniya s poiskom podkhodyashchego akkaunta"""
        try:
            with self.lock:
                wb = self._load_workbook_safe()
                ws = wb.active

            accounts = []
            total_quota_collected = 0
            processed_count = 0
            changes_made = False
            
            log.info(f"📊 Scanning Excel rows for reassignment (need {required_quota} chars)")
            
            for row in range(2, ws.max_row + 1):
                api_key = ws[f'A{row}'].value
                email = ws[f'B{row}'].value
                status = ws[f'F{row}'].value or 'active'
                unusual_activity = ws[f'I{row}'].value or 'no'

                if not api_key:
                    continue
                    
                processed_count += 1
                log.debug(f"📋 Row {row}: {email}, status={status}, unusual={unusual_activity}")

                if (status and status.lower() == 'disabled') or unusual_activity == 'yes':
                    log.debug(f"⏭️ Skipping {email}: status={status}, unusual={unusual_activity}")
                    continue
                
                last_checked = ws[f'E{row}'].value
                quota = ws[f'D{row}'].value

                # Obnovlyaem kvotu esli ona pustaya, ravna 0, ravna None ili ustarela
                should_check_quota = (
                    quota is None or 
                    quota == 0 or 
                    quota == ''
                    # self._is_quota_check_needed(last_checked)  # SPRAVLENE: ispolzuem self
                )
                
                if should_check_quota:
                    reason = 'empty' if quota is None or quota == 0 or quota == '' else 'outdated'
                    log.info(f"📊 Checking quota for reassignment account: {email} (reason: {reason})")

                    if self.mobile_proxy:
                        proxy_info = self.mobile_proxy.get_proxy_connection_info()
                        if proxy_info:
                            proxy_dict = {
                                "http": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                                "https": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}"
                            }
                        else:
                            log.error("❌ No proxy connection info for quota check")
                            continue
                    else:
                        log.error("❌ No mobile proxy available for quota check")
                        continue

                    quota = self.check_quota(api_key, proxy_dict, force=True)

                    
                    # SOKhRANYaEM KVOTU SRAZU
                    ws[f'D{row}'] = quota
                    ws[f'E{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    changes_made = True
                    
                    # SOKhRANYaEM FAYL POSLE KAZhDOGO OBNOVLENYa
                    wb.save(self.excel_path)
                    log.info(f"💾 Saved reassignment quota for {email}: {quota}")
                    
                    time.sleep(2)  # Pauza mezhdu zaprosami

                # Dobavlyaem akkaunt v spisok esli u nego est kvota (bolshe 0)
                if quota and quota > 0:
                    accounts.append({
                        'api_key': api_key,
                        'email': email,
                        'quota_remaining': quota,
                        'row': row
                    })
                    total_quota_collected += quota
                    log.info(f"✅ Reassignment account {email}: {quota} characters available")
                    
                    # Ostanavlivaemsya tolko esli nabrana dostatochnaya obshchaya kvota
                    # i est khotya by odin akkaunt, kotoryy pokryvaet trebuemuyu kvotu
                    if required_quota > 0 and total_quota_collected >= required_quota:
                        if any(acc['quota_remaining'] >= required_quota for acc in accounts):
                            log.info(
                                f"🎯 Found sufficient reassignment quota: {total_quota_collected} >= {required_quota}"
                            )
                            break
                else:
                    log.warning(f"⚠️ Reassignment account {email}: no quota available (quota={quota})")
            
            # Finalnoe sokhranenie esli byli izmeneniya
            if changes_made:
                wb.save(self.excel_path)
                
            wb.close()
            
            # Sortiruem po ubyvaniyu kvoty
            accounts.sort(key=lambda x: x['quota_remaining'], reverse=True)
            
            log.info(f"📊 For reassignment: processed {processed_count} rows, found {len(accounts)} accounts with total quota: {total_quota_collected}")
            
            return accounts
            
        except Exception as e:
            log.error(f"❌ Error getting all available accounts for reassignment: {e}")
            return []

    def _process_reassigned_requests(self):
        """Obrabatyvaet perenaznachennye zaprosy"""
        log.info("🔄 Starting processing of reassigned requests")
        # spolzuem sushchestvuyushchuyu logiku obrabotki akkauntov
        self._process_accounts_sequentially()
        
    def _make_elevenlabs_request(self, account: dict, request: dict) -> dict:
        """Odin zapros k ElevenLabs API s uchetom:
        - globalnogo semafora na akkaunt (<=2 odnovremennykh);
        - retraev setevykh oshibok.
        """
        api_key = account['api_key']
        sem = self._get_account_semaphore(api_key)
        sem.acquire()

        try:
            if self.stop_event.is_set():
                return {'success': False, 'error': 'Shutdown in progress'}

            attempt = 0
            max_attempts = 3
            while True:
                if self.stop_event.is_set():
                    return {'success': False, 'error': 'Shutdown in progress'}

                attempt += 1
                proxy_dict = None
                if self.mobile_proxy:
                    proxy_info = self.mobile_proxy.get_proxy_connection_info()
                    if not proxy_info:
                        return {'success': False, 'error': 'Mobile proxy connection failed'}
                    proxy_dict = {
                        "http":  f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                        "https": f"http://{proxy_info['username']}:{proxy_info['password']}@{proxy_info['host']}:{proxy_info['port']}",
                    }
                else:
                    return {'success': False, 'error': 'No mobile proxy available'}

                # Ensure voices cleaned once before first request for this account
                self._ensure_initial_voice_cleanup(account, proxy_dict)

                session = requests.Session()
                session.trust_env = False
                session.proxies = proxy_dict

                url = f"https://api.elevenlabs.io/v1/text-to-speech/{request['voice_id']}"
                headers = {
                    'Accept': 'audio/mpeg',
                    'Content-Type': 'application/json',
                    'xi-api-key': api_key,
                }

                model_id = request['config'].get('model_id', self.config.get('model_id'))
                allowed = MODEL_VOICE_PARAMS.get(model_id, VOICE_DEFAULTS.keys())
                voice_settings = {
                    p: request['config'].get(p, VOICE_DEFAULTS.get(p))
                    for p in allowed
                    if request['config'].get(p, VOICE_DEFAULTS.get(p)) is not None
                }
                payload = {
                    'text': request['text'],
                    'model_id': model_id,
                    'voice_settings': voice_settings,
                }

                log.info("📤 ElevenLabs Request: POST %s", url)
                log.info("📤 Account: %s", account['email'])
                headers_log = {
                    k: (v if FULL_LOGS or k != 'xi-api-key' else f"{v[:10]}...")
                    for k, v in headers.items()
                }
                log.info("📤 Headers: %s", json.dumps(headers_log, indent=2))
                payload_log = {**payload, 'text': maybe_truncate(request['text'], 100)}
                log.info(
                    "📤 Payload: %s",
                    json.dumps(payload_log, indent=2, ensure_ascii=False),
                )

                response = None
                duration = None
                try:
                    start = time.time()
                    response = session.post(
                        url, json=payload, headers=headers,
                    )
                    duration = time.time() - start
                except (ReadTimeout, ConnectionError, socket.error) as e:
                    try:
                        session.close()
                    except Exception:
                        pass
                    log.warning(f"⏳ Request attempt {attempt} failed: {e}")
                    time.sleep(5)
                    continue
                finally:
                    try:
                        session.close()
                    except Exception:
                        pass

                log.info("📥 ElevenLabs Response: %d (took %.2fs)", response.status_code, duration)
                try:
                    log.info("📥 Response Headers: %s", json.dumps(dict(response.headers), indent=2))
                except Exception:
                    pass

                if response.status_code == 200:
                    log.info("📥 Response Size: %d bytes", len(response.content))
                    return {'success': True, 'content': response.content, 'content_type': 'audio/mpeg'}

                try:
                    log.info("📥 Error Response Body: %s", maybe_truncate(response.text, 500))
                except Exception:
                    pass

                error_status = 'unknown'
                error_detail = {}
                try:
                    error_data = response.json()
                    error_detail = error_data.get('detail', {})
                    error_status = error_detail.get('status', 'unknown')
                except Exception:
                    pass

                if error_status == 'quota_exceeded':
                    message = error_detail.get('message', '')
                    remaining = 0
                    m = re.search(r'You have (\\d+) credits remaining', message or '')
                    if m:
                        remaining = int(m.group(1))
                    self.mark_quota_exceeded(api_key, remaining, message)
                    return {'success': False, 'error': 'quota_exceeded', 'remaining': remaining}

                if error_status == 'voice_limit_reached':
                    log.warning("Voice limit reached; cleaning up custom voices and retrying")
                    try:
                        log.info("📥 Response Size: %d bytes", len(response.content))
                        self.ensure_account_voices_cleaned(api_key, account.get('email'), proxy_dict)
                    except Exception:
                        pass
                    if attempt == max_attempts:
                        return {'success': False, 'error': 'voice_limit_reached'}
                    time.sleep(2)
                    continue

                if error_status == 'detected_unusual_activity':
                    log.warning("🚨 Suspicious activity detected for %s (attempt %d)", account['email'], attempt)
                    return {'success': False, 'error': 'suspicious_activity'}

                return {'success': False, 'error': f'API key error: {error_status}'}

                # 429 — slishkom mnogo odnovremennykh zaprosov: podozhdem i povtorim (no semafor i tak derzhit <=2)
                if response.status_code == 429:
                    log.warning("⚠️ HTTP 429: too many concurrent requests; backing off and retrying...")
                    # Dop. zashchita: nebolshaya pauza
                    time.sleep(3)
                    continue

                # Prochie oshibki
                return {'success': False, 'error': f'HTTP {response.status_code}: {response.text}'}

            return {'success': False, 'error': 'Request failed'}

        finally:
            try:
                sem.release()
            except Exception:
                pass

    def _process_accounts_concurrently(self):
        """Parallelno obrabatyvaet akkaunty, ogranichivaya chislo odnovremenno rabotayushchikh akkauntov."""
        suspicious_requests = []
        while self.account_assignments and not self.stop_event.is_set() and not self.quota_refresh_needed:
            assignments_items = list(self.account_assignments.items())

            for i in range(0, len(assignments_items), self.max_parallel_accounts):
                if self.stop_event.is_set() or self.quota_refresh_needed:
                    break

                batch = assignments_items[i : i + self.max_parallel_accounts]
                threads = []

                for account_id, assignment in batch:
                    if self.stop_event.is_set() or self.quota_refresh_needed:
                        break
                    if account_id not in self.account_assignments:
                        continue

                    account = assignment["account"]
                    requests = assignment["requests"]
                    if not requests:
                        log.info(
                            f"🎤 Skipping account {account['email']}: no requests to process"
                        )
                        self.account_assignments.pop(account_id, None)
                        continue

                    log.info(
                        f"🎤 Processing account {account['email']}: {len(requests)} requests"
                    )

                    suspicious_holder = []

                    def worker(acc=account, reqs=requests, holder=suspicious_holder):
                        if self.stop_event.is_set() or self.quota_refresh_needed:
                            return
                        self._process_remaining_requests_fast(acc, reqs)
                        for r in reqs:
                            if r.get("status") == "suspicious":
                                holder.append(r)

                    thread = threading.Thread(target=worker, daemon=True)
                    threads.append((thread, suspicious_holder, account_id))
                    thread.start()

                for thread, holder, acc_id in threads:
                    thread.join()
                    suspicious_requests.extend(holder)
                    self.account_assignments.pop(acc_id, None)

        if self.quota_refresh_needed:
            remaining = []
            for assignment in self.account_assignments.values():
                for req in assignment["requests"]:
                    if req.get("status") != "completed":
                        req["status"] = "pending"
                        remaining.append(req)
            self.account_assignments.clear()
            # Obnovlyaem kvoty tolko u nuzhnykh akkauntov
            self.refresh_all_quotas(
                accounts=list(self.quota_refresh_accounts)
            )
            self.quota_refresh_accounts.clear()
            self.quota_refresh_needed = False
            for req in remaining:
                self.queue.put(req)
            log.info(f"🔄 Requeued {len(remaining)} requests after quota refresh")

        return suspicious_requests

    def _process_accounts_concurrently(self):
        suspicious_requests = []
        lock = threading.Lock()

        # Если есть глобальный семафор (например, для лимита batch_size) — используем его.
        global_sem = getattr(self, "_global_semaphore", None)

        while self.account_assignments and not self.stop_event.is_set() and not self.quota_refresh_needed:
            # Sobiraem vse pary (account, request)
            work_items = []
            for account_id, assignment in list(self.account_assignments.items()):
                account = assignment.get("account")
                for req in assignment.get("requests", []):
                    if req.get("status") not in ("completed", "failed"):
                        work_items.append((account_id, account, req))

            if not work_items:
                self.account_assignments.clear()
                break

            threads = []

            def worker(acc, req):
                if self.stop_event.is_set() or self.quota_refresh_needed:
                    return
                self._process_single_request_with_quota_update(acc, req)
                if req.get("status") == "suspicious":
                    with lock:
                        suspicious_requests.append(req)

            # Zapuskaem po potoku na kazhdyy zapros
            for _acc_id, acc, req in work_items:
                t = threading.Thread(target=worker, args=(acc, req), daemon=True)
                threads.append((_acc_id, t))
                t.start()

            for _acc_id, t in threads:
                try:
                    t.join()
                except Exception as exc:  # на всякий случай, чтобы не прерывать весь цикл
                    log.error(f"⚠️ Не удалось дождаться потока для аккаунта {_acc_id}: {exc}")

            for acc_id in list(self.account_assignments.keys()):
                assignment = self.account_assignments.get(acc_id)
                if not assignment:
                    continue
                # ostavlyaem tolko te, chto eshche ne zaversheny
                assignment["requests"] = [r for r in assignment["requests"] if r.get("status") not in ("completed", "failed")]

                if not assignment["requests"]:
                    self.account_assignments.pop(acc_id, None)

        # Если во время обработки решили обновить квоты — перекидываем незавершённые запросы в очередь
        if self.quota_refresh_needed:
            remaining = []
            for assignment in self.account_assignments.values():
                for req in assignment.get("requests", []):
                    if req.get("status") != "completed":
                        req["status"] = "pending"
                        remaining.append(req)
            self.account_assignments.clear()

            try:
                self.refresh_all_quotas(accounts=list(self.quota_refresh_accounts))
            except Exception as exc:
                log.error(f"❌ Ошибка обновления квот: {exc}")

            self.quota_refresh_accounts.clear()
            self.quota_refresh_needed = False

            for req in remaining:
                self.queue.put(req)
            log.info(f"🔄 Requeued {len(remaining)} requests after quota refresh")

        return suspicious_requests


    def _process_remaining_requests_fast(self, account: dict, remaining_requests: list):
        """Bystro obrabatyvaet ostalnye zaprosy s ogranicheniem parallelizma"""
        max_concurrent = self.max_concurrent_per_account
        log.info(
            f"🚀 Starting LIMITED PARALLEL processing of {len(remaining_requests)} requests (max_concurrent={max_concurrent}) for {account['email']}"
        )

        # Razbivaem na batchi
        import threading

        semaphore = threading.Semaphore(max_concurrent)
        global_sem = getattr(self, "_global_semaphore", None)  # может быть None

        def worker(req):
            if self.stop_event.is_set() or self.quota_refresh_needed:
                return
            # Небольшой джиттер, чтобы не выстрелить пачкой в одну миллисекунду
            time.sleep(random.uniform(0.5, 1.5))
            if self.stop_event.is_set() or self.quota_refresh_needed:
                return

            with semaphore:
                if self.stop_event.is_set() or self.quota_refresh_needed:
                    return
                # Если задан глобальный семафор — учитываем и его
                if global_sem is not None:
                    with global_sem:
                        if self.stop_event.is_set() or self.quota_refresh_needed:
                            return
                        try:
                            self._process_single_request_with_quota_update(account, req)
                        except Exception as e:
                            log.error(f"❌ Thread error for req {req.get('id')}: {e}", exc_info=True)
                else:
                    try:
                        self._process_single_request_with_quota_update(account, req)
                    except Exception as e:
                        log.error(f"❌ Thread error for req {req['id']}: {e}", exc_info=True)
        
        # Sozdaem potoki s ogranicheniem
        threads = []
        for req in remaining_requests:
            if self.stop_event.is_set():
                break
            if self.quota_refresh_needed:
                break
            thread = threading.Thread(target=process_with_semaphore, args=(req,), daemon=True)
            threads.append(thread)
            thread.start()
            time.sleep(0.5)  # Zaderzhka mezhdu zapuskami potokov

        # Zhdem zaversheniya
        for thread in threads:
            thread.join()

    def _process_single_request_with_quota_update(self, account: dict, req: dict):
        """Obrabatyvaet odin zapros s obnovleniem kvoty (dlya parallelnogo vypolneniya)"""
        if self.stop_event.is_set():
            log.info(f"⏹️ Skipping request {req['id']} due to shutdown")
            return

        log.info(f"🎵 Processing request {req['id']} for {account['email']} in parallel")

        if self.stop_event.is_set():
            log.info(f"⏹️ Aborted request {req['id']} before start")
            return

        result = self._make_elevenlabs_request(account, req)
        req['result'] = result

        if result['success']:
            req['status'] = 'completed'
            self._store_result(req['id'], result)
            chars_needed = len(req['text'])
            model_id = req['config'].get('model_id', self.config.get('model_id'))
            self.update_quota_after_request(account['api_key'], chars_needed, model_id)
            log.info(f"✅ Request {req['id']} completed successfully, quota updated")
        elif result.get('error') == 'suspicious_activity':
            req['status'] = 'suspicious'
            log.warning(f"🚨 Suspicious activity for request {req['id']} on {account['email']}")
        elif 'quota_exceeded' in str(result.get('error')):
            log.warning(
                f"⚠️ Quota exceeded for {account['email']}, reassigning request"
            )
            req['status'] = 'pending'
            req['result'] = None
            with self.lock:
                assignment = self.account_assignments.get(account['api_key'])
                if assignment and req in assignment['requests']:
                    assignment['requests'].remove(req)
                    if not assignment['requests']:
                        self.account_assignments.pop(account['api_key'], None)
            self._reassign_failed_requests([req], account['email'])
        else:
            req['status'] = 'failed'
            self._store_result(req['id'], result)
            log.error(
                f"❌ Request {req['id']} failed: {result.get('error', 'Unknown error')}"
            )

    def _rotate_ip_for_account(self, email: str) -> bool:
        """Rotiruet IP pered nachalom raboty s akkauntom. Vozvrashchaet True/False.
        Esli rotatsiya/validnyy IP ne polucheny za otvedennoe vremya — NE prodolzhaem rabotu s etim akkauntom.
        """
        if not self.mobile_proxy:
            log.error("❌ No mobile proxy available for rotation")
            return False

        max_total_wait_s = 180      # obshchiy limit ozhidaniya rotatsii
        attempt_sleep_s  = 5        # pauza mezhdu popytkami
        start_ts = time.time()

        log.info(f"🔄 Rotating IP for account: {email}")

        while time.time() - start_ts < max_total_wait_s:
            try:
                ok = self.mobile_proxy.rotate_ip()
                if not ok:
                    # menedzher sam proveryaet status rotatsii; zhdem i snova
                    remain = max_total_wait_s - (time.time() - start_ts)
                    log.warning(f"⚠️ IP rotation attempt failed; retrying in {attempt_sleep_s:.0f}s "
                                f"(time left ~{int(remain)}s)")
                    time.sleep(attempt_sleep_s)
                    continue

                # proverim, chto IP deystvitelno est/validnyy
                new_ip = self.mobile_proxy.get_current_ip()
                if not new_ip or str(new_ip).lower() == 'unknown':
                    log.warning("⚠️ Rotation returned unknown IP; retrying...")
                    time.sleep(attempt_sleep_s)
                    continue

                log.info(f"✅ IP rotated for {email}: {new_ip}")
                log.info("⏳ Waiting 5 seconds for connection to stabilize...")
                time.sleep(5)
                return True

            except Exception as e:
                remain = max_total_wait_s - (time.time() - start_ts)
                log.error(f"❌ Exception during IP rotation for {email}: {e}. Retrying in {attempt_sleep_s}s "
                          f"(time left ~{int(remain)}s)")
                time.sleep(attempt_sleep_s)

        log.error(f"❌ IP rotation HARD timeout for {email} (>{max_total_wait_s}s). Aborting this account.")
        return False


    def _test_account_with_first_request(self, account: dict, first_request: dict) -> bool:
        """Testiruet akkaunt pervym zaprosom s proverkoy kvoty i proksi/IP.
        Vozvrashchaet True — esli zapros uspeshen i mozhno prodolzhat na etom akkaunte.
        """
        max_attempts = 3

        # Proverka kvoty do lyubykh setevykh deystviy
        current_quota = self.check_and_update_quota_from_excel(account['api_key'])
        chars_needed = len(first_request['text'])
        if current_quota < chars_needed:
            log.warning(f"⚠️ Insufficient quota for test request: need {chars_needed}, have {current_quota}")
            return False

        # Ensure voices cleaned before first use of this account
        self._ensure_initial_voice_cleanup(account, None)
            
        for attempt in range(1, max_attempts + 1):
            log.info(f"🧪 Testing account {account['email']}, attempt {attempt}/{max_attempts} (quota: {current_quota})")

            # Proveryaem realnyy IP proksi. Esli neizvesten — probuem fors-rotatsiyu.
            if self.mobile_proxy:
                try:
                    ip = self.mobile_proxy.get_current_ip()
                    log.info(f"📡 Current proxy IP before request: {ip}")
                    if not ip or str(ip).lower() == 'unknown':
                        log.warning("⚠️ Proxy IP unknown. Forcing rotation before test...")
                        if not self._rotate_ip_for_account(account['email']):
                            log.error("❌ Rotation failed; aborting this account")
                            return False
                except Exception as e:
                    log.warning(f"⚠️ Could not verify proxy IP: {e}. Forcing rotation...")
                    if not self._rotate_ip_for_account(account['email']):
                        log.error("❌ Rotation failed; aborting this account")
                        return False

            # Sam zapros
            result = self._make_elevenlabs_request(account, first_request)

            if result['success']:
                log.info(f"✅ Account {account['email']} test successful")
                first_request['result'] = result
                first_request['status'] = 'completed'
                self._store_result(first_request['id'], result)

                model_id = first_request['config'].get('model_id', self.config.get('model_id'))
                self.update_quota_after_request(account['api_key'], chars_needed, model_id)
                return True

            # Yavnyy perekhvat prevysheniya kvoty
            if 'quota_exceeded' in str(result.get('error')):
                log.warning(
                    f"⚠️ Quota exceeded during test for {account['email']}, reassigning request"
                )
                first_request['status'] = 'pending'
                first_request['result'] = None
                self._reassign_failed_requests([first_request], account['email'])
                return False

            # Unusual activity — ne krutim beskonechno, prosto pomechaem i vykhodim
            if 'unusual' in (result.get('error', '').lower()):
                log.warning(f"🚨 Unusual activity on {account['email']}")
                self.mark_unusual_activity(account['api_key'], account['email'], 1)
                return False

            # Oshibki soedineniya — odna rotatsiya i povtor
            err_msg = result.get('error', 'Unknown error')
            if attempt < max_attempts and ('proxy' in err_msg.lower() or 'connection' in err_msg.lower()):
                log.error(f"❌ Proxy/connection error for {account['email']}: {err_msg}")
                if not self._rotate_ip_for_account(account['email']):
                    log.error("❌ Rotation failed; aborting this account")
                    return False
                time.sleep(5)
                continue

            log.error(f"❌ Account {account['email']} failed: {err_msg}")
            return False

        return False

